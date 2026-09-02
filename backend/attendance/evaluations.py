"""Evaluation forms: what a student may answer, and what the answers add up to.

The college's evaluations were collected on paper and tallied by hand, so the
only thing anyone ever saw was the tally — never the distribution, never a
comparison against last year, and never the free-text answers unless somebody
retyped them. Holding the questions as data makes all three fall out for free.

Every rule about who may answer what, and what an answer is allowed to look
like, lives here so the portal, the API and the export cannot disagree.
"""

from collections import Counter
from datetime import date

from django.db import transaction

from .models import (
    AcademicYear, Form, FormAnswer, FormQuestion, FormResponse, FormSubmissionReceipt,
)

TEXT_TYPES = {FormQuestion.SHORT_TEXT, FormQuestion.LONG_TEXT}
CHOICE_TYPES = {FormQuestion.SINGLE_CHOICE, FormQuestion.MULTI_CHOICE}

MAX_SHORT_TEXT = 300
MAX_LONG_TEXT = 5000


# ── what a student can see ────────────────────────────────────────────────────

#: "do not filter by level at all", which is a different thing from a student
#: whose level we could not work out — that one gets no level-specific forms.
ANY_LEVEL = object()


def open_forms(today=None, audience=Form.STUDENT, class_level=ANY_LEVEL, kind=None):
    """Every form this audience may answer right now.

    Defaults to the student portal, because a staff-completed form appearing in
    a student's list would invite them to evaluate themselves. Pass a
    `class_level` to drop the forms aimed at some other NTA level, and a `kind`
    to separate the evaluations from the things a student is asking for.
    """
    today = today or date.today()
    qs = (Form.objects.filter(is_active=True)
          .select_related('academic_year').prefetch_related('levels'))
    if audience is not None:
        qs = qs.filter(audience=audience)
    if kind is not None:
        qs = qs.filter(kind=kind)
    forms = [form for form in qs if form.is_open(today)]
    if class_level is not ANY_LEVEL:
        forms = [form for form in forms if form.applies_to(class_level)]
    return forms


def level_of(profile):
    """The NTA level whose forms this student is asked for.

    Derived here rather than asked of the caller: a caller that forgets would
    quietly widen a form to every level, and nothing would look wrong.

    Scoped to the active year, because a continuing student is enrolled at
    level 4 last year and level 5 this one, and it is this year's level the
    college is asking about.
    """
    if profile is None:
        return None
    from . import finance          # local — finance owns enrollment → level
    year = AcademicYear.objects.filter(is_active=True).first()
    return finance.class_level_for(profile, year)


def answered_form_ids(profile):
    if profile is None:
        return set()
    return set(
        FormSubmissionReceipt.objects.filter(profile=profile).values_list('form_id', flat=True)
    )


def pending_mandatory_forms(profile, today=None):
    """The forms this student must answer before the portal will let them past.

    Read off the submission receipts, not the responses, so a mandatory form
    can still be anonymous — the college learns that everyone has answered
    without learning who said what.
    """
    if profile is None:
        return []
    answered = answered_form_ids(profile)
    # Evaluations only. A service request cannot be made compulsory — asking a
    # student for a sick sheet they do not need is nonsense — and filtering here
    # means bad data cannot lock the portal behind one either.
    pending = [form for form in open_forms(today, class_level=level_of(profile),
                                           kind=Form.EVALUATION)
               if form.is_mandatory and form.id not in answered]
    # Form's own ordering is newest-first, which would ask for the most recent
    # census before one that has been waiting a fortnight. Oldest first.
    pending.sort(key=lambda form: (form.created_at, form.id))
    return pending


def forms_for_student(profile, today=None, kind=None):
    """The open forms, each marked with whether this student has answered it.

    Answered forms are still listed rather than hidden — a student who has
    filled one in wants to see that it is done, not wonder whether it saved.
    """
    answered = answered_form_ids(profile)
    return [
        {'form': form, 'answered': form.id in answered,
         'can_answer': form.allow_multiple or form.id not in answered}
        for form in open_forms(today, class_level=level_of(profile), kind=kind)
    ]


# ── services a student asks for ───────────────────────────────────────────────

def services_for_student(profile, today=None):
    """The services this student may request — a sick sheet, a letter, a bed."""
    return forms_for_student(profile, today, kind=Form.REQUEST)


def my_requests(profile):
    """This student's own service requests, newest first, with where each got to.

    A request is the one thing on the portal the college owes an answer to, so
    the student sees the decision and the note that came with it — where to
    collect the letter, or why it was turned down.
    """
    if profile is None:
        return []
    return list(
        FormResponse.objects
        .filter(form__kind=Form.REQUEST, profile=profile)
        .select_related('form')
        .order_by('-submitted_at')
    )


def request_queue(status=None):
    """Every service request the college has been sent, newest first."""
    qs = (FormResponse.objects
          .filter(form__kind=Form.REQUEST)
          .select_related('form', 'profile', 'class_level', 'decided_by')
          .prefetch_related('answers__question')
          .order_by('-submitted_at'))
    return qs.filter(status=status) if status else qs


def decide(response, *, status, note='', by=None):
    """Answer a service request.

    Sending one back to pending is allowed and clears the decision with it —
    an officer who approved the wrong request should be able to undo it, not
    live with a decision nobody made.
    """
    from django.utils import timezone

    if response.form.kind != Form.REQUEST:
        raise ValueError('Only a service request can be approved or declined.')
    if status not in dict(FormResponse.STATUS_CHOICES):
        raise ValueError('That is not a decision.')

    response.status = status
    response.decision_note = (note or '').strip()
    decided = status != FormResponse.PENDING
    response.decided_by = by if decided else None
    response.decided_at = timezone.now() if decided else None
    response.save(update_fields=['status', 'decision_note', 'decided_by', 'decided_at'])
    return response


# ── validating an answer ──────────────────────────────────────────────────────

def answerable_questions(form):
    """The questions the student is actually asked.

    A section marked `for_office` belongs to a health facility or a signing
    officer and is printed blank on the approved document, so it is not part of
    the form the student fills in — nor of anything counted, summarised or
    exported, where it would only ever be an empty column.
    """
    return FormQuestion.objects.filter(section__form=form, section__for_office=False)


def _clean_text(value, limit):
    text = '' if value is None else str(value).strip()
    if len(text) > limit:
        raise ValueError(f'Answer is too long (max {limit} characters).')
    return text


def normalise_answer(question, value):
    """Coerce a submitted answer into the shape its question stores.

    Raises ValueError with a message meant for the student, not the developer.
    Returns None when nothing was answered, so the caller can decide whether
    that is allowed.
    """
    label = question.text[:60]

    if question.type == FormQuestion.SHORT_TEXT:
        return _clean_text(value, MAX_SHORT_TEXT) or None

    if question.type == FormQuestion.LONG_TEXT:
        return _clean_text(value, MAX_LONG_TEXT) or None

    if question.type == FormQuestion.SINGLE_CHOICE:
        choice = _clean_text(value, MAX_SHORT_TEXT)
        if not choice:
            return None
        if choice not in question.options:
            raise ValueError(f'"{choice}" is not one of the choices for "{label}".')
        return choice

    if question.type == FormQuestion.MULTI_CHOICE:
        if value in (None, ''):
            return None
        if not isinstance(value, (list, tuple)):
            raise ValueError(f'"{label}" expects a list of choices.')
        chosen = [_clean_text(item, MAX_SHORT_TEXT) for item in value]
        chosen = [item for item in chosen if item]
        for item in chosen:
            if item not in question.options:
                raise ValueError(f'"{item}" is not one of the choices for "{label}".')
        return chosen or None

    if question.type == FormQuestion.MATRIX:
        if not value:
            return None
        if not isinstance(value, dict):
            raise ValueError(f'"{label}" expects a rating for each row.')
        ratings = {}
        for row, choice in value.items():
            row = _clean_text(row, MAX_SHORT_TEXT)
            choice = _clean_text(choice, MAX_SHORT_TEXT)
            if not choice:
                continue
            if row not in question.rows:
                raise ValueError(f'"{row}" is not a row of "{label}".')
            if choice not in question.options:
                raise ValueError(f'"{choice}" is not a rating for "{label}".')
            ratings[row] = choice
        return ratings or None

    if question.type == FormQuestion.GRID_TEXT:
        if not value:
            return None
        if not isinstance(value, (list, tuple)):
            raise ValueError(f'"{label}" expects a table of rows.')
        width = len(question.columns) or 1
        table = []
        for raw_row in value[:question.max_rows or len(value)]:
            if isinstance(raw_row, dict):
                cells = [raw_row.get(column, '') for column in question.columns]
            elif isinstance(raw_row, (list, tuple)):
                cells = list(raw_row)
            else:
                cells = [raw_row]
            cells = [_clean_text(cell, MAX_SHORT_TEXT) for cell in cells[:width]]
            cells += [''] * (width - len(cells))
            if any(cells):                      # drop the rows nobody filled in
                table.append(cells)
        return table or None

    raise ValueError(f'Unknown question type: {question.type}')


# ── submitting ────────────────────────────────────────────────────────────────

@transaction.atomic
def submit(form, answers, *, profile=None, class_level=None, academic_year=None,
           submitted_by=None, today=None):
    """Record one filled-in form.

    `answers` is {question_id: value}. Everything is validated before anything
    is written, so a form is never half-saved.

    On an anonymous form the response is stored with no profile at all, and the
    receipt that stops a second submission is written separately — there is no
    join that would put the two back together.
    """
    if not form.is_open(today):
        raise ValueError('That form is not open for responses.')

    # A form aimed at another NTA level is refused here as well as hidden from
    # the list, so knowing the slug is not a way in.
    if profile is not None and not form.applies_to(class_level or level_of(profile)):
        raise ValueError('That form is not open for responses.')

    if profile is not None and not form.allow_multiple:
        if FormSubmissionReceipt.objects.filter(form=form, profile=profile).exists():
            raise ValueError('You have already answered this form.')

    questions = list(answerable_questions(form).select_related('section'))
    by_id = {question.id: question for question in questions}

    # Keys arrive as JSON object keys, so they are strings. A key that is not a
    # number at all used to raise int()'s own message, which was then shown to
    # the student as "invalid literal for int() with base 10".
    try:
        submitted = {int(key) for key in answers}
    except (TypeError, ValueError):
        raise ValueError('That answer does not belong to this form.')
    if submitted - set(by_id):
        raise ValueError('That answer does not belong to this form.')

    cleaned = {}
    for question in questions:
        raw = answers.get(question.id, answers.get(str(question.id)))
        value = normalise_answer(question, raw)
        if value is None:
            if question.required:
                raise ValueError(f'"{question.text[:80]}" must be answered.')
            continue
        cleaned[question.id] = value

    response = FormResponse.objects.create(
        form=form,
        profile=None if form.is_anonymous else profile,
        class_level=class_level,
        academic_year=academic_year or form.academic_year,
        # Never on an anonymous form: recording the member of staff who filled
        # one in would identify the respondent just as surely as a name.
        submitted_by=None if form.is_anonymous else submitted_by,
    )
    FormAnswer.objects.bulk_create([
        FormAnswer(response=response, question_id=question_id, value=value)
        for question_id, value in cleaned.items()
    ])
    if profile is not None:
        FormSubmissionReceipt.objects.get_or_create(form=form, profile=profile)
    return response


# ── what the answers add up to ────────────────────────────────────────────────

def _mean(pairs):
    """Weighted mean of [(numeric value, count)], to one decimal."""
    total = sum(count for _value, count in pairs)
    if not total:
        return None
    return round(sum(value * count for value, count in pairs) / total, 1)


def _choice_summary(question, values):
    counts = Counter(values)
    numeric = question.numeric_options
    scale = dict(zip(question.options, numeric)) if numeric else {}
    return {
        'counts': [{'option': option, 'count': counts.get(option, 0)}
                   for option in question.options],
        'answered': sum(counts.values()),
        'mean': _mean([(scale[option], counts.get(option, 0)) for option in question.options])
        if scale else None,
    }


def summarise_question(question, answers):
    """One question's results, in the shape a chart and a table both want."""
    values = [answer.value for answer in answers]
    base = {
        'id': question.id,
        'text': question.text,
        'type': question.type,
        'section': question.section.title,
        'answered': len(values),
    }

    if question.type in TEXT_TYPES:
        return {**base, 'kind': 'text',
                'responses': [value for value in values if value]}

    if question.type == FormQuestion.SINGLE_CHOICE:
        return {**base, 'kind': 'choice', **_choice_summary(question, values)}

    if question.type == FormQuestion.MULTI_CHOICE:
        flat = [item for value in values if isinstance(value, list) for item in value]
        return {**base, 'kind': 'choice', **_choice_summary(question, flat)}

    if question.type == FormQuestion.MATRIX:
        numeric = question.numeric_options
        scale = dict(zip(question.options, numeric)) if numeric else {}
        rows = []
        for row in question.rows:
            counts = Counter(
                value[row] for value in values
                if isinstance(value, dict) and value.get(row)
            )
            rows.append({
                'row': row,
                'counts': [{'option': option, 'count': counts.get(option, 0)}
                           for option in question.options],
                'answered': sum(counts.values()),
                'mean': _mean([(scale[option], counts.get(option, 0))
                               for option in question.options]) if scale else None,
            })
        overall = [row['mean'] for row in rows if row['mean'] is not None]
        return {**base, 'kind': 'matrix', 'options': question.options, 'rows': rows,
                'mean': round(sum(overall) / len(overall), 1) if overall else None}

    # grid_text — a table of free text; there is nothing to average, so the
    # rows people actually wrote are the answer.
    table = [row for value in values if isinstance(value, list) for row in value]
    return {**base, 'kind': 'table', 'columns': question.columns, 'rows': table}


def summarise(form):
    """Every question's results, ready for the charts and the summary sheet."""
    questions = list(
        answerable_questions(form)
        .select_related('section')
        .order_by('section__order', 'order', 'id')
    )
    answers = {}
    for answer in FormAnswer.objects.filter(response__form=form).select_related('question'):
        answers.setdefault(answer.question_id, []).append(answer)

    return {
        'form': {'id': form.id, 'title': form.title, 'slug': form.slug,
                 'status': form.status(), 'is_anonymous': form.is_anonymous},
        'responses': FormResponse.objects.filter(form=form).count(),
        'questions': [summarise_question(q, answers.get(q.id, [])) for q in questions],
    }


# ── the spreadsheet ───────────────────────────────────────────────────────────

def _flatten(question, value):
    """One answer as the cells it occupies in the response sheet."""
    if value is None:
        return [''] * len(export_columns(question))
    if question.type == FormQuestion.MATRIX:
        return [value.get(row, '') if isinstance(value, dict) else '' for row in question.rows]
    if question.type == FormQuestion.MULTI_CHOICE:
        return ['; '.join(value) if isinstance(value, list) else str(value)]
    if question.type == FormQuestion.GRID_TEXT:
        rows = [' | '.join(str(cell) for cell in row) for row in value] \
            if isinstance(value, list) else [str(value)]
        return ['\n'.join(rows)]
    return [str(value)]


def answer_text(question, value):
    """One answer as a line of prose, for a printed document.

    The spreadsheet wants an answer split across cells; a printed form wants it
    read back as a sentence, so a rating table becomes "Row — rating" and a
    table of text becomes one row per line.
    """
    if value is None or value == '':
        return ''
    if isinstance(value, dict):
        return '; '.join(f'{row} — {value[row]}' for row in question.rows if row in value)
    if isinstance(value, list):
        if value and isinstance(value[0], (list, tuple)):
            return '\n'.join(' | '.join(str(cell) for cell in row) for row in value)
        return '; '.join(str(item) for item in value)
    return str(value)


def export_columns(question):
    """A rating table needs one column per row rated; everything else needs one."""
    if question.type == FormQuestion.MATRIX:
        return [f'{question.text[:60]} — {row}' for row in question.rows]
    return [question.text[:120]]


def export_workbook(form):
    """The form's responses as a workbook: every answer, then the tallies.

    Two sheets rather than one because they answer different questions —
    "what did people say" and "what does it add up to" — and the college
    reports from the second while checking against the first.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', fgColor='1E2D78')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

    questions = list(
        answerable_questions(form)
        .select_related('section').order_by('section__order', 'order', 'id')
    )
    responses = list(
        FormResponse.objects.filter(form=form)
        .select_related('profile', 'class_level', 'academic_year', 'submitted_by')
        .prefetch_related('answers')
        .order_by('submitted_at')
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Responses'

    identity = ['#', 'Submitted']
    if not form.is_anonymous:
        identity += (['Filled in by'] if form.audience == form.STAFF
                     else ['Student', 'Registration no.'])
    identity += ['Level', 'Academic year']

    header = list(identity)
    for question in questions:
        header += export_columns(question)
    sheet.append(header)
    for cell in sheet[1]:
        cell.fill, cell.font, cell.alignment = header_fill, header_font, wrap

    for index, response in enumerate(responses, start=1):
        answers = {answer.question_id: answer.value for answer in response.answers.all()}
        row = [index, response.submitted_at.strftime('%d %b %Y %H:%M')]
        if not form.is_anonymous:
            if form.audience == form.STAFF:
                staff = response.submitted_by
                row += [(staff.get_full_name() or staff.username) if staff else '—']
            else:
                row += [response.profile.name if response.profile else '—',
                        response.profile.nactvet_reg_no if response.profile else '—']
        row += [response.class_level.name if response.class_level else '—',
                response.academic_year.name if response.academic_year else '—']
        for question in questions:
            row += _flatten(question, answers.get(question.id))
        sheet.append(row)

    sheet.freeze_panes = 'A2'
    for column in range(1, len(header) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 22 if column > len(identity) else 14

    # ── the tally ──
    tally = workbook.create_sheet('Summary')
    tally.append(['Question', 'Row', 'Answer', 'Count', 'Average'])
    for cell in tally[1]:
        cell.fill, cell.font, cell.alignment = header_fill, header_font, wrap

    data = summarise(form)
    for question in data['questions']:
        if question['kind'] == 'choice':
            for entry in question['counts']:
                tally.append([question['text'], '', entry['option'], entry['count'],
                              question['mean'] if question['mean'] is not None else ''])
        elif question['kind'] == 'matrix':
            for row in question['rows']:
                for entry in row['counts']:
                    tally.append([question['text'], row['row'], entry['option'],
                                  entry['count'],
                                  row['mean'] if row['mean'] is not None else ''])
        else:
            tally.append([question['text'], '', 'Free text', question['answered'], ''])

    for column, width in zip('ABCDE', (54, 34, 24, 10, 10)):
        tally.column_dimensions[column].width = width
    tally.freeze_panes = 'A2'
    return workbook
