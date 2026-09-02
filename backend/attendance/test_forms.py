"""Evaluation forms: who may answer, what an answer may be, and what it adds up to.

The college's evaluations were paper, tallied by hand. The point of holding them
as data is that the tally stops being the only thing anyone sees — so these
tests care as much about the summary and the export as about the submission.
"""

from datetime import date, timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import Client, TestCase
from rest_framework.test import APIClient

from . import evaluations
from .serializers import FormResponseSerializer
from .models import (
    AcademicYear, ClassLevel, Form, FormAnswer, FormQuestion, FormResponse,
    FormSection, FormSubmissionReceipt, Module, SecretaryProfile, Semester, Student,
    TeacherProfile,
)

User = get_user_model()


class FormTestBase(TestCase):
    def setUp(self):
        self.year = AcademicYear.objects.create(name='2026/2027', is_active=True)
        self.sem = Semester.objects.create(academic_year=self.year, number=1, is_active=True)
        self.level = ClassLevel.objects.create(name='NTA Level 4', order=4)
        self.module = Module.objects.create(name='Pharmaceutics', code='PHM101', teacher='T',
                                            class_level=self.level, semester=self.sem)
        self.enrollment = Student.objects.create(
            nactvet_reg_no='BPH/2026/001', name='Asha Juma', module=self.module)
        self.enrollment.set_portal_pin('Portal#2026', require_change=False)
        self.enrollment.save()
        from . import finance
        self.asha = finance.profile_for_student(self.enrollment)

        self.admin = User.objects.create_superuser('admin', 'a@b.c', 'pw')
        self.api = APIClient()
        self.api.force_authenticate(self.admin)

        self.portal = Client()
        self.portal.post('/login/', {'identifier': 'BPH/2026/001', 'secret': 'Portal#2026'})

    def _form(self, **kw):
        defaults = {'title': 'Course Evaluation', 'slug': 'course-eval', 'is_active': True}
        defaults.update(kw)
        form = Form.objects.create(**defaults)
        self.section = FormSection.objects.create(form=form, title='Section A', order=0)
        return form

    def _question(self, **kw):
        defaults = {'section': self.section, 'type': FormQuestion.SINGLE_CHOICE,
                    'text': 'Rate the tutor', 'options': ['Excellent', 'Good', 'Poor']}
        defaults.update(kw)
        return FormQuestion.objects.create(**defaults)


class WhatStudentsCanSeeTests(FormTestBase):
    """Only active forms, and only inside their window."""

    def test_a_student_sees_only_active_forms(self):
        self._form(title='Live one', slug='live', is_active=True)
        self._form(title='Draft one', slug='draft', is_active=False)

        listed = self.portal.get('/api/my-forms/').json()
        self.assertEqual([f['slug'] for f in listed], ['live'])

    def test_a_form_is_hidden_before_it_opens_and_after_it_closes(self):
        today = date(2026, 10, 15)
        early = self._form(slug='early', opens_on=today + timedelta(days=3))
        late = self._form(slug='late', closes_on=today - timedelta(days=1))
        now = self._form(slug='now', opens_on=today - timedelta(days=1),
                         closes_on=today + timedelta(days=1))

        self.assertEqual(early.status(today), Form.DRAFT)
        self.assertEqual(late.status(today), Form.CLOSED)
        self.assertEqual(now.status(today), Form.OPEN)
        self.assertEqual([f.slug for f in evaluations.open_forms(today)], ['now'])

    def test_an_inactive_form_cannot_be_opened_or_answered_directly(self):
        form = self._form(slug='draft', is_active=False)
        self._question()
        self.assertEqual(self.portal.get(f'/api/my-forms/{form.slug}/').status_code, 404)

        response = self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                                    {'answers': {}}, content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn('not open', response.json()['detail'])

    def test_a_signed_out_visitor_gets_nothing(self):
        self._form()
        self.assertEqual(Client().get('/api/my-forms/').status_code, 403)


class AnsweringTests(FormTestBase):
    def test_answers_are_recorded_against_the_form(self):
        form = self._form()
        rating = self._question(required=True)
        comment = self._question(type=FormQuestion.LONG_TEXT, text='Comments', options=[])

        response = self.portal.post(
            f'/api/my-forms/{form.slug}/submit/',
            {'answers': {str(rating.id): 'Good', str(comment.id): 'The labs were excellent.'}},
            content_type='application/json')
        self.assertEqual(response.status_code, 201, response.content)

        saved = FormResponse.objects.get(form=form)
        self.assertEqual(saved.profile, self.asha)
        self.assertEqual(saved.class_level, self.level)
        self.assertEqual(saved.academic_year, self.year)
        self.assertEqual(
            {a.question_id: a.value for a in saved.answers.all()},
            {rating.id: 'Good', comment.id: 'The labs were excellent.'})

    def test_a_required_question_must_be_answered(self):
        form = self._form()
        self._question(required=True, text='Rate the field practice')

        response = self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                                    {'answers': {}}, content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn('must be answered', response.json()['detail'])
        self.assertFalse(FormResponse.objects.exists())

    def test_an_answer_outside_the_choices_is_refused(self):
        form = self._form()
        rating = self._question()

        response = self.portal.post(
            f'/api/my-forms/{form.slug}/submit/',
            {'answers': {str(rating.id): 'Superb'}}, content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertIn('not one of the choices', response.json()['detail'])

    def test_a_malformed_answer_key_is_refused_cleanly(self):
        """A junk key used to surface int()'s own message to the student."""
        form = self._form()
        self._question()
        response = self.portal.post(
            f'/api/my-forms/{form.slug}/submit/', {'answers': {'not-a-number': 'Good'}},
            content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['detail'],
                         'That answer does not belong to this form.')
        self.assertFalse(FormResponse.objects.exists())

    def test_nothing_is_saved_when_one_answer_is_bad(self):
        """A form is never half-recorded."""
        form = self._form()
        good = self._question(text='Rate the tutor')
        bad = self._question(text='Rate the labs')

        self.portal.post(
            f'/api/my-forms/{form.slug}/submit/',
            {'answers': {str(good.id): 'Good', str(bad.id): 'Nonsense'}},
            content_type='application/json')
        self.assertFalse(FormResponse.objects.exists())
        self.assertFalse(FormAnswer.objects.exists())

    def test_a_student_answers_once(self):
        form = self._form()
        rating = self._question()
        payload = {'answers': {str(rating.id): 'Good'}}

        first = self.portal.post(f'/api/my-forms/{form.slug}/submit/', payload,
                                 content_type='application/json')
        self.assertEqual(first.status_code, 201)
        second = self.portal.post(f'/api/my-forms/{form.slug}/submit/', payload,
                                  content_type='application/json')
        self.assertEqual(second.status_code, 400)
        self.assertIn('already answered', second.json()['detail'])
        self.assertEqual(FormResponse.objects.count(), 1)

    def test_a_form_can_allow_more_than_one_response(self):
        """One response per tutor, so the tutor evaluation must repeat."""
        form = self._form(allow_multiple=True)
        rating = self._question()
        payload = {'answers': {str(rating.id): 'Good'}}

        for _ in range(3):
            self.assertEqual(
                self.portal.post(f'/api/my-forms/{form.slug}/submit/', payload,
                                 content_type='application/json').status_code, 201)
        self.assertEqual(FormResponse.objects.count(), 3)

    def test_the_form_lists_what_i_have_already_answered(self):
        form = self._form()
        rating = self._question()
        listed = self.portal.get('/api/my-forms/').json()[0]
        self.assertFalse(listed['answered'])
        self.assertTrue(listed['can_answer'])

        self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                         {'answers': {str(rating.id): 'Good'}},
                         content_type='application/json')
        listed = self.portal.get('/api/my-forms/').json()[0]
        self.assertTrue(listed['answered'])
        self.assertFalse(listed['can_answer'])


class AnonymityTests(FormTestBase):
    """A student will not say a tutor was unprepared with their name on it."""

    def test_an_anonymous_response_carries_no_link_to_the_student(self):
        form = self._form(is_anonymous=True)
        rating = self._question()

        self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                         {'answers': {str(rating.id): 'Poor'}},
                         content_type='application/json')

        saved = FormResponse.objects.get(form=form)
        self.assertIsNone(saved.profile)
        # ...and nothing the admin can read off it leads back to her either.
        self.assertNotIn(self.asha.nactvet_reg_no,
                         str(FormResponseSerializer(saved).data))

    def test_an_anonymous_form_still_knows_who_has_answered(self):
        form = self._form(is_anonymous=True)
        rating = self._question()
        self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                         {'answers': {str(rating.id): 'Poor'}},
                         content_type='application/json')

        receipt = FormSubmissionReceipt.objects.get(form=form)
        self.assertEqual(receipt.profile, self.asha)
        # Which is how a second attempt is refused without de-anonymising anyone.
        again = self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                                 {'answers': {str(rating.id): 'Good'}},
                                 content_type='application/json')
        self.assertEqual(again.status_code, 400)

    def test_the_admin_can_chase_non_responders_without_reading_names_off_answers(self):
        form = self._form(is_anonymous=True)
        rating = self._question()
        self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                         {'answers': {str(rating.id): 'Poor'}},
                         content_type='application/json')

        listed = self.api.get(f'/api/forms/{form.id}/who-responded/').data
        self.assertEqual(listed['count'], 1)
        self.assertEqual(listed['answered'][0]['reg_no'], 'BPH/2026/001')

        # The responses themselves name nobody.
        responses = self.api.get(f'/api/forms/{form.id}/responses/').data
        self.assertEqual(responses[0]['student_name'], 'Anonymous')
        self.assertEqual(responses[0]['student_reg_no'], '')


class RatingTableTests(FormTestBase):
    ROWS = ['Water facility', 'Electricity facility', 'Safety and Security']

    def _matrix_form(self):
        form = self._form(slug='hostel')
        question = self._question(
            type=FormQuestion.MATRIX, text='Hostel facilities',
            options=['1', '2', '3', '4', '5'], rows=self.ROWS)
        return form, question

    def test_a_rating_table_is_stored_row_by_row(self):
        form, question = self._matrix_form()
        self.portal.post(
            f'/api/my-forms/{form.slug}/submit/',
            {'answers': {str(question.id): {'Water facility': '2',
                                            'Electricity facility': '4'}}},
            content_type='application/json')

        answer = FormAnswer.objects.get(question=question)
        self.assertEqual(answer.value, {'Water facility': '2', 'Electricity facility': '4'})

    def test_a_rating_outside_the_scale_or_an_unknown_row_is_refused(self):
        form, question = self._matrix_form()
        for bad in ({'Water facility': '9'}, {'Swimming pool': '3'}):
            response = self.portal.post(
                f'/api/my-forms/{form.slug}/submit/', {'answers': {str(question.id): bad}},
                content_type='application/json')
            self.assertEqual(response.status_code, 400, bad)
        self.assertFalse(FormResponse.objects.exists())

    def test_a_table_of_text_drops_the_rows_nobody_filled_in(self):
        form = self._form(slug='grid')
        question = self._question(
            type=FormQuestion.GRID_TEXT, text='Modules with poor materials',
            options=[], columns=['Module name', 'What was wrong'], max_rows=4)

        self.portal.post(
            f'/api/my-forms/{form.slug}/submit/',
            {'answers': {str(question.id): [['PHM101', 'No notes'], ['', ''], ['ANA101', '']]}},
            content_type='application/json')

        answer = FormAnswer.objects.get(question=question)
        self.assertEqual(answer.value, [['PHM101', 'No notes'], ['ANA101', '']])


class SummaryTests(FormTestBase):
    def _answer(self, reg_no, question, value):
        enrollment = Student.objects.create(
            nactvet_reg_no=reg_no, name=f'Student {reg_no}', module=self.module)
        enrollment.set_portal_pin('Portal#2026', require_change=False)
        enrollment.save()
        from . import finance
        profile = finance.profile_for_student(enrollment)
        evaluations.submit(question.section.form, {question.id: value}, profile=profile,
                           class_level=self.level, academic_year=self.year)

    def test_a_choice_question_is_counted_per_option(self):
        form = self._form()
        question = self._question(options=['Excellent', 'Good', 'Poor'])
        for index, value in enumerate(['Good', 'Good', 'Poor']):
            self._answer(f'BPH/S/{index}', question, value)

        summary = evaluations.summarise(form)
        self.assertEqual(summary['responses'], 3)
        counts = summary['questions'][0]['counts']
        self.assertEqual(counts, [{'option': 'Excellent', 'count': 0},
                                  {'option': 'Good', 'count': 2},
                                  {'option': 'Poor', 'count': 1}])
        # Words do not average — only a numeric scale does.
        self.assertIsNone(summary['questions'][0]['mean'])

    def test_a_numeric_scale_averages(self):
        form = self._form()
        question = self._question(options=['1', '2', '3', '4', '5'])
        for index, value in enumerate(['5', '4', '3']):
            self._answer(f'BPH/N/{index}', question, value)

        self.assertEqual(evaluations.summarise(form)['questions'][0]['mean'], 4.0)

    def test_a_rating_table_averages_each_row_and_overall(self):
        form = self._form(slug='hostel')
        question = self._question(type=FormQuestion.MATRIX, text='Hostel facilities',
                                  options=['1', '2', '3', '4', '5'],
                                  rows=['Water facility', 'Electricity facility'])
        self._answer('BPH/M/1', question, {'Water facility': '2', 'Electricity facility': '4'})
        self._answer('BPH/M/2', question, {'Water facility': '4', 'Electricity facility': '4'})

        summary = evaluations.summarise(form)['questions'][0]
        self.assertEqual(summary['kind'], 'matrix')
        rows = {row['row']: row for row in summary['rows']}
        self.assertEqual(rows['Water facility']['mean'], 3.0)
        self.assertEqual(rows['Electricity facility']['mean'], 4.0)
        self.assertEqual(summary['mean'], 3.5)

    def test_free_text_answers_are_listed_rather_than_counted(self):
        form = self._form()
        question = self._question(type=FormQuestion.LONG_TEXT, text='Comments', options=[])
        self._answer('BPH/T/1', question, 'More practicals please.')

        summary = evaluations.summarise(form)['questions'][0]
        self.assertEqual(summary['kind'], 'text')
        self.assertEqual(summary['responses'], ['More practicals please.'])

    def test_the_summary_endpoint_serves_it(self):
        form = self._form()
        question = self._question()
        self._answer('BPH/E/1', question, 'Good')

        data = self.api.get(f'/api/forms/{form.id}/summary/').data
        self.assertEqual(data['responses'], 1)
        self.assertEqual(data['questions'][0]['text'], question.text)


class ExportTests(FormTestBase):
    def test_the_workbook_has_the_answers_and_the_tally(self):
        form = self._form()
        rating = self._question(text='Rate the tutor')
        table = self._question(type=FormQuestion.MATRIX, text='Hostel facilities',
                               options=['1', '2', '3', '4', '5'],
                               rows=['Water facility', 'Electricity facility'])
        evaluations.submit(form, {rating.id: 'Good',
                                  table.id: {'Water facility': '2'}},
                           profile=self.asha, class_level=self.level, academic_year=self.year)

        workbook = evaluations.export_workbook(form)
        self.assertEqual(workbook.sheetnames, ['Responses', 'Summary'])

        sheet = workbook['Responses']
        header = [cell.value for cell in sheet[1]]
        self.assertIn('Student', header)
        self.assertIn('Rate the tutor', header)
        # A rating table takes one column per row rated.
        self.assertIn('Hostel facilities — Water facility', header)
        self.assertIn('Hostel facilities — Electricity facility', header)

        row = [cell.value for cell in sheet[2]]
        self.assertIn('Asha Juma', row)
        self.assertIn('Good', row)
        self.assertIn('2', row)

    def test_an_anonymous_export_has_no_name_columns(self):
        form = self._form(is_anonymous=True)
        rating = self._question()
        evaluations.submit(form, {rating.id: 'Poor'}, profile=self.asha)

        sheet = evaluations.export_workbook(form)['Responses']
        header = [cell.value for cell in sheet[1]]
        self.assertNotIn('Student', header)
        self.assertNotIn('Registration no.', header)
        self.assertNotIn('Asha Juma', [cell.value for cell in sheet[2]])

    def test_the_download_endpoint_serves_a_spreadsheet(self):
        form = self._form()
        self._question()
        response = self.api.get(f'/api/forms/{form.id}/export/')
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])
        self.assertIn(f'{form.slug}_responses.xlsx', response['Content-Disposition'])

        from openpyxl import load_workbook
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['Responses', 'Summary'])


class AdminAccessTests(FormTestBase):
    def test_only_the_admin_can_publish_a_form(self):
        tutor = User.objects.create_user('tutor', password='pw')
        TeacherProfile.objects.create(user=tutor, full_name='Tutor')
        api = APIClient()
        api.force_authenticate(tutor)

        self.assertEqual(api.post('/api/forms/', {'title': 'Mine'}, format='json').status_code, 403)
        # A tutor may still read them.
        self.assertEqual(api.get('/api/forms/').status_code, 200)

    def test_a_form_is_created_with_a_slug_of_its_own(self):
        first = self.api.post('/api/forms/', {'title': 'Course Evaluation'}, format='json')
        second = self.api.post('/api/forms/', {'title': 'Course Evaluation'}, format='json')
        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(first.data['slug'], 'course-evaluation')
        self.assertEqual(second.data['slug'], 'course-evaluation-2')

    def test_a_closing_date_before_the_opening_date_is_refused(self):
        response = self.api.post('/api/forms/', {
            'title': 'Backwards', 'opens_on': '2026-10-10', 'closes_on': '2026-10-01',
        }, format='json')
        self.assertEqual(response.status_code, 400)

    def test_a_question_that_cannot_be_answered_is_refused(self):
        self._form()
        for payload in (
            {'section': self.section.id, 'text': 'Pick one', 'type': FormQuestion.SINGLE_CHOICE},
            {'section': self.section.id, 'text': 'Rate', 'type': FormQuestion.MATRIX,
             'options': ['1', '2']},
        ):
            response = self.api.post('/api/form-questions/', payload, format='json')
            self.assertEqual(response.status_code, 400, payload)

    def test_a_question_with_answers_cannot_be_deleted(self):
        form = self._form()
        rating = self._question()
        evaluations.submit(form, {rating.id: 'Good'}, profile=self.asha)

        response = self.api.delete(f'/api/form-questions/{rating.id}/')
        self.assertEqual(response.status_code, 403)


class SeededFormsTests(TestCase):
    """The college's four real forms, loaded from its own documents."""

    def test_the_colleges_forms_load_and_match_the_paper(self):
        call_command('seed_evaluation_forms', verbosity=0)
        by_slug = {form.slug: form for form in Form.objects.all()}
        self.assertEqual(sorted(by_slug), [
            'course-evaluation', 'hostel-accommodation-request', 'hostel-evaluation',
            'mentorship-evaluation', 'official-letter-request', 'practicum-letter-request',
            'practicum-letter-request-level-6', 'sick-sheet-request',
            'student-performance-evaluation', 'student-permission-request',
            'tracer-study', 'tutor-evaluation'])

        # The things a student asks for are requests, not evaluations: they are
        # owed an answer, and they appear under Services on the portal.
        self.assertEqual(
            sorted(form.slug for form in by_slug.values() if form.kind == Form.REQUEST),
            ['hostel-accommodation-request', 'official-letter-request',
             'practicum-letter-request', 'practicum-letter-request-level-6',
             'sick-sheet-request', 'student-permission-request'])
        # None of them anonymous — the college has to know whose request it is.
        self.assertFalse(any(form.is_anonymous for form in by_slug.values()
                             if form.kind == Form.REQUEST))

        # Anything evaluating a member of staff is anonymous.
        self.assertTrue(by_slug['tutor-evaluation'].is_anonymous)
        self.assertTrue(by_slug['hostel-evaluation'].is_anonymous)
        self.assertFalse(by_slug['course-evaluation'].is_anonymous)

        # They arrive unpublished, so the admin checks them over first.
        self.assertFalse(any(form.is_active for form in by_slug.values()))

        hostel = FormQuestion.objects.get(section__form=by_slug['hostel-evaluation'],
                                          type=FormQuestion.MATRIX)
        self.assertEqual(hostel.options, ['1', '2', '3', '4', '5'])
        self.assertEqual(len(hostel.rows), 12)
        self.assertIn('Patron/Matron services', hostel.rows)

        tutor = FormQuestion.objects.get(section__form=by_slug['tutor-evaluation'],
                                         type=FormQuestion.MATRIX)
        self.assertEqual(len(tutor.rows), 15)

        offices = FormQuestion.objects.filter(
            section__form=by_slug['course-evaluation'], type=FormQuestion.MATRIX).first()
        self.assertEqual(len(offices.rows), 15)

    def test_the_mentorship_form_matches_its_document(self):
        call_command('seed_evaluation_forms', verbosity=0)
        form = Form.objects.get(slug='mentorship-evaluation')
        # The paper marks the name Optional — the college asking people to
        # speak freely — so the form is anonymous rather than nearly so.
        self.assertTrue(form.is_anonymous)
        self.assertEqual(form.audience, Form.STUDENT)
        self.assertEqual(form.sections.count(), 5)

        outcomes = FormQuestion.objects.get(section__form=form, type=FormQuestion.MATRIX)
        self.assertEqual(outcomes.options,
                         ['Strongly Disagree', 'Disagree', 'Agree', 'Strongly Agree'])
        self.assertEqual(len(outcomes.rows), 4)

        role = FormQuestion.objects.get(section__form=form, text='Role')
        self.assertEqual(role.options, ['Mentor', 'Mentee'])

    def test_the_performance_evaluation_is_filled_in_by_staff_not_students(self):
        """A mentor assesses a student with it. If it reached the student's own
        list they would be grading themselves."""
        call_command('seed_evaluation_forms', activate=True, verbosity=0)
        form = Form.objects.get(slug='student-performance-evaluation')
        self.assertEqual(form.audience, Form.STAFF)
        self.assertTrue(form.allow_multiple)     # one per student assessed

        self.assertNotIn(form, evaluations.open_forms())
        self.assertIn(form, evaluations.open_forms(audience=Form.STAFF))

        # Its five rating tables use the paper's Always → N/A scale.
        tables = FormQuestion.objects.filter(section__form=form, type=FormQuestion.MATRIX)
        self.assertEqual(tables.count(), 5)
        for table in tables:
            self.assertEqual(table.options,
                             ['Always', 'Sometimes', 'Seldom', 'Never', 'N/A'])
        self.assertEqual(sum(len(t.rows) for t in tables), 28)

    def test_the_practicum_request_captures_the_students_half_of_the_form(self):
        call_command('seed_evaluation_forms', verbosity=0)
        form = Form.objects.get(slug='practicum-letter-request')
        self.assertEqual(form.audience, Form.STUDENT)
        self.assertFalse(form.is_anonymous)      # the office must know who asked
        self.assertTrue(form.allow_multiple)

        # Parts A to C only — D and E are the college's approval chain.
        self.assertEqual([s.title.split(':')[0] for s in form.sections.all()],
                         ['Part A', 'Part B', 'Part C'])
        self.assertIn('completed by the college after you submit', form.intro)

        declaration = FormQuestion.objects.get(section__form=form, options=['I agree'])
        self.assertTrue(declaration.required)

    def test_re_running_it_does_not_disturb_answered_forms(self):
        call_command('seed_evaluation_forms', verbosity=0)
        form = Form.objects.get(slug='hostel-evaluation')
        question = FormQuestion.objects.filter(section__form=form).first()
        FormAnswer.objects.create(
            response=FormResponse.objects.create(form=form), question=question, value={})

        call_command('seed_evaluation_forms', verbosity=0)
        self.assertTrue(FormQuestion.objects.filter(id=question.id).exists())

    def test_activating_publishes_them(self):
        call_command('seed_evaluation_forms', activate=True, verbosity=0)
        self.assertTrue(all(form.is_active for form in Form.objects.all()))


class AudienceTests(FormTestBase):
    """A staff-completed form must never reach a student."""

    def test_a_staff_form_is_absent_from_the_students_list(self):
        self._form(title='Mentor assessment', slug='mentor-assess', audience=Form.STAFF)
        self._question()
        self.assertEqual(self.portal.get('/api/my-forms/').json(), [])

    def test_a_student_cannot_open_or_submit_a_staff_form_by_its_slug(self):
        form = self._form(slug='mentor-assess', audience=Form.STAFF)
        rating = self._question()

        self.assertEqual(self.portal.get(f'/api/my-forms/{form.slug}/').status_code, 404)
        blocked = self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                                   {'answers': {str(rating.id): 'Good'}},
                                   content_type='application/json')
        self.assertEqual(blocked.status_code, 404)
        self.assertFalse(FormResponse.objects.exists())

    def test_the_admin_still_sees_and_exports_staff_forms(self):
        form = self._form(slug='mentor-assess', audience=Form.STAFF)
        self._question()
        listed = self.api.get('/api/forms/').data
        self.assertIn(form.id, [f['id'] for f in listed])
        self.assertEqual(self.api.get(f'/api/forms/{form.id}/export/').status_code, 200)

    def test_forms_default_to_the_student_portal(self):
        created = self.api.post('/api/forms/', {'title': 'New one'}, format='json')
        self.assertEqual(created.data['audience'], Form.STUDENT)


class DeletingAFormTests(FormTestBase):
    """A form the college has stopped running must be removable — but not by
    accident, because deleting one takes its responses with it."""

    def test_an_unanswered_form_deletes_straight_away(self):
        form = self._form()
        self._question()
        self.assertEqual(self.api.delete(f'/api/forms/{form.id}/').status_code, 204)
        self.assertFalse(Form.objects.filter(id=form.id).exists())
        self.assertFalse(FormQuestion.objects.exists())

    def test_an_answered_form_asks_before_destroying_the_answers(self):
        form = self._form()
        rating = self._question()
        evaluations.submit(form, {rating.id: 'Good'}, profile=self.asha)

        refused = self.api.delete(f'/api/forms/{form.id}/')
        self.assertEqual(refused.status_code, 409)
        self.assertTrue(refused.data['requires_confirmation'])
        self.assertEqual(refused.data['responses'], 1)
        self.assertIn('Download the Excel first', refused.data['detail'])
        self.assertTrue(Form.objects.filter(id=form.id).exists())

    def test_confirming_deletes_the_form_and_its_answers(self):
        form = self._form()
        rating = self._question()
        evaluations.submit(form, {rating.id: 'Good'}, profile=self.asha)

        self.assertEqual(
            self.api.delete(f'/api/forms/{form.id}/?confirm=yes').status_code, 204)
        self.assertFalse(Form.objects.filter(id=form.id).exists())
        self.assertFalse(FormResponse.objects.exists())
        self.assertFalse(FormAnswer.objects.exists())

    def test_the_deletion_is_audited(self):
        form = self._form(title='Old evaluation')
        rating = self._question()
        evaluations.submit(form, {rating.id: 'Good'}, profile=self.asha)
        self.api.delete(f'/api/forms/{form.id}/?confirm=yes')

        from .models import FinanceAuditLog
        entry = FinanceAuditLog.objects.filter(action='form.delete').first()
        self.assertIsNotNone(entry)
        self.assertEqual(entry.actor, self.admin)
        self.assertIn('Old evaluation', entry.summary)
        self.assertEqual(entry.before['responses'], 1)

    def test_a_tutor_cannot_delete_a_form(self):
        form = self._form()
        tutor = User.objects.create_user('tutor2', password='pw')
        TeacherProfile.objects.create(user=tutor, full_name='Tutor')
        api = APIClient()
        api.force_authenticate(tutor)
        self.assertEqual(api.delete(f'/api/forms/{form.id}/').status_code, 403)
        self.assertTrue(Form.objects.filter(id=form.id).exists())

    def test_a_section_with_answers_is_protected(self):
        form = self._form()
        rating = self._question()
        evaluations.submit(form, {rating.id: 'Good'}, profile=self.asha)

        refused = self.api.delete(f'/api/form-sections/{self.section.id}/')
        self.assertEqual(refused.status_code, 403)
        self.assertIn('delete the whole form', refused.data['detail'])

    def test_an_unanswered_section_deletes_with_its_questions(self):
        self._form()
        self._question()
        self.assertEqual(
            self.api.delete(f'/api/form-sections/{self.section.id}/').status_code, 204)
        self.assertFalse(FormQuestion.objects.exists())


class StaffFilledFormTests(FormTestBase):
    """The Students' Performance Evaluation is a mentor assessing a student, so
    it needs somewhere to be filled in that is not the student's own portal."""

    def _staff_form(self):
        form = self._form(slug='performance', audience=Form.STAFF, allow_multiple=True)
        question = self._question(type=FormQuestion.MATRIX, text='Leadership',
                                  options=['Always', 'Sometimes', 'Never'],
                                  rows=['Able to cooperate with others'])
        name = self._question(type=FormQuestion.SHORT_TEXT, text='Students Name',
                              options=[], required=True)
        return form, question, name

    def test_the_admin_can_fill_a_staff_form_in(self):
        form, matrix, name = self._staff_form()
        response = self.api.post(f'/api/forms/{form.id}/submit/', {'answers': {
            str(name.id): 'Asha Juma',
            str(matrix.id): {'Able to cooperate with others': 'Always'},
        }}, format='json')
        self.assertEqual(response.status_code, 201, response.data)

        saved = FormResponse.objects.get(form=form)
        self.assertEqual(saved.submitted_by, self.admin)
        self.assertIsNone(saved.profile)
        self.assertEqual(saved.answers.count(), 2)

    def test_a_mentor_can_assess_more_than_one_student(self):
        form, matrix, name = self._staff_form()
        for student in ('Asha Juma', 'Baraka Simon'):
            self.assertEqual(self.api.post(f'/api/forms/{form.id}/submit/', {'answers': {
                str(name.id): student,
                str(matrix.id): {'Able to cooperate with others': 'Always'},
            }}, format='json').status_code, 201)
        self.assertEqual(FormResponse.objects.count(), 2)

    def test_a_required_question_is_enforced_here_too(self):
        form, matrix, _name = self._staff_form()
        refused = self.api.post(f'/api/forms/{form.id}/submit/', {'answers': {
            str(matrix.id): {'Able to cooperate with others': 'Always'},
        }}, format='json')
        self.assertEqual(refused.status_code, 400)
        self.assertIn('must be answered', refused.data['detail'])

    def test_a_student_form_cannot_be_filled_in_from_the_panel(self):
        form = self._form(slug='course', audience=Form.STUDENT)
        rating = self._question()
        refused = self.api.post(f'/api/forms/{form.id}/submit/',
                                {'answers': {str(rating.id): 'Good'}}, format='json')
        self.assertEqual(refused.status_code, 400)
        self.assertIn('answered by students on their portal', refused.data['detail'])

    def test_a_tutor_cannot_fill_a_staff_form_in(self):
        form, matrix, name = self._staff_form()
        tutor = User.objects.create_user('tutor3', password='pw')
        TeacherProfile.objects.create(user=tutor, full_name='Tutor')
        api = APIClient()
        api.force_authenticate(tutor)
        self.assertEqual(api.post(f'/api/forms/{form.id}/submit/',
                                  {'answers': {}}, format='json').status_code, 403)

    def test_the_export_names_who_filled_it_in(self):
        form, matrix, name = self._staff_form()
        self.api.post(f'/api/forms/{form.id}/submit/', {'answers': {
            str(name.id): 'Asha Juma',
            str(matrix.id): {'Able to cooperate with others': 'Always'},
        }}, format='json')

        sheet = evaluations.export_workbook(form)['Responses']
        header = [cell.value for cell in sheet[1]]
        self.assertIn('Filled in by', header)
        self.assertNotIn('Registration no.', header)
        self.assertIn('Asha Juma', [cell.value for cell in sheet[2]])

    def test_an_anonymous_form_never_records_the_member_of_staff(self):
        """Recording who filled an anonymous form in would identify the
        respondent just as surely as a name."""
        form = self._form(slug='anon-staff', audience=Form.STAFF, is_anonymous=True)
        rating = self._question()
        self.api.post(f'/api/forms/{form.id}/submit/',
                      {'answers': {str(rating.id): 'Good'}}, format='json')
        saved = FormResponse.objects.get(form=form)
        self.assertIsNone(saved.submitted_by)
        self.assertIsNone(saved.profile)


class EditingAQuestionTests(FormTestBase):
    def test_a_question_can_be_corrected_after_it_is_answered(self):
        """A typo used to mean deleting and retyping, which is refused once
        anyone has answered."""
        form = self._form()
        rating = self._question(text='Rate the tutar')
        evaluations.submit(form, {rating.id: 'Good'}, profile=self.asha)

        response = self.api.patch(f'/api/form-questions/{rating.id}/', {
            'text': 'Rate the tutor',
            'options': ['Excellent', 'Good', 'Poor'],
            'help_text': 'Think about the whole year',
        }, format='json')
        self.assertEqual(response.status_code, 200, response.data)

        rating.refresh_from_db()
        self.assertEqual(rating.text, 'Rate the tutor')
        self.assertEqual(rating.help_text, 'Think about the whole year')
        # The answer already given is untouched.
        self.assertEqual(FormAnswer.objects.get(question=rating).value, 'Good')

    def test_an_edit_that_would_orphan_an_answer_is_still_checked(self):
        form = self._form()
        rating = self._question()
        refused = self.api.patch(f'/api/form-questions/{rating.id}/',
                                 {'options': []}, format='json')
        self.assertEqual(refused.status_code, 400)


class StudentResultsNavigationTests(TestCase):
    """The results area is two screens: CA for the semester in progress, and
    every published end-of-semester result.

    CA is a running figure, not a result. Once the college advances the
    semester the marks are superseded by the end-of-semester result, so they
    stop being shown — but nothing published is ever lost, because the
    end-of-semester screen keeps it.
    """

    def setUp(self):
        from .models import StudentResult
        self.year = AcademicYear.objects.create(name='2026/2027', is_active=True)
        self.sem1 = Semester.objects.create(academic_year=self.year, number=1, is_active=True)
        self.sem2 = Semester.objects.create(academic_year=self.year, number=2, is_active=False)
        self.level = ClassLevel.objects.create(name='NTA Level 4', order=4)
        self.mod1 = Module.objects.create(name='Pharmaceutics', code='PHM101', teacher='T',
                                          class_level=self.level, semester=self.sem1)
        self.mod2 = Module.objects.create(name='Anatomy', code='ANA101', teacher='T',
                                          class_level=self.level, semester=self.sem2)
        for module in (self.mod1, self.mod2):
            enrollment = Student.objects.create(
                nactvet_reg_no='BPH/2026/001', name='Asha Juma', module=module)
            enrollment.set_portal_pin('Portal#2026', require_change=False)
            enrollment.save()
            StudentResult.objects.create(
                student=enrollment, assign1=30, assign2=30, cat1_theory=60, cat2_theory=60,
                ca_approved=True)
        self.portal = Client()
        self.portal.post('/login/', {'identifier': 'BPH/2026/001', 'secret': 'Portal#2026'})

    def _context(self):
        return self.portal.get('/student-dashboard/').context

    def test_the_results_menu_offers_exactly_two_screens(self):
        body = self.portal.get('/student-dashboard/').content.decode()
        self.assertIn('data-view="ca-results"', body)
        self.assertIn('data-view="final-results"', body)
        # The per-semester CA screens are gone, and What I Owe has moved to
        # Payments where the rest of the money lives.
        self.assertNotIn('data-view="ca-sem1"', body)
        self.assertNotIn('data-view="ca-sem2"', body)
        self.assertIn('id="payments-menu"', body)

    def test_ca_shows_only_the_semester_in_progress(self):
        context = self._context()
        codes = [m['module_code'] for m in context['ca_theory_modules']]
        self.assertEqual(codes, ['PHM101'])          # semester 1 is active
        self.assertTrue(context['has_ca_results'])

    def test_advancing_the_semester_takes_the_old_ca_away(self):
        """Exactly what the college sees after the admin advances."""
        self.assertEqual(len(self._context()['ca_theory_modules']), 1)

        self.sem1.is_active = False
        self.sem1.save(update_fields=['is_active'])
        self.sem2.is_active = True
        self.sem2.save(update_fields=['is_active'])

        context = self._context()
        codes = [m['module_code'] for m in context['ca_theory_modules']]
        self.assertEqual(codes, ['ANA101'])          # semester 2 now
        self.assertNotIn('PHM101', codes)

    def test_ca_empties_when_the_new_semester_has_no_marks_yet(self):
        from .models import StudentResult
        StudentResult.objects.filter(student__module=self.mod2).update(ca_approved=False)
        self.sem1.is_active = False
        self.sem1.save(update_fields=['is_active'])
        self.sem2.is_active = True
        self.sem2.save(update_fields=['is_active'])

        context = self._context()
        self.assertFalse(context['has_ca_results'])
        self.assertEqual(context['ca_theory_modules'], [])
        self.assertContains(self.portal.get('/student-dashboard/'),
                            'No CA results have been published for')

    def test_end_of_semester_keeps_every_published_semester(self):
        from .models import StudentResult
        # Publish both semesters, then advance past semester 1.
        StudentResult.objects.all().update(
            end_theory=70, final_approved=True, ca_approved=True)
        self.sem1.is_active = False
        self.sem1.save(update_fields=['is_active'])
        self.sem2.is_active = True
        self.sem2.save(update_fields=['is_active'])

        context = self._context()
        published = [m['module_code'] for m in context['modules'] if m['has_final_result']]
        self.assertCountEqual(published, ['PHM101', 'ANA101'])
        self.assertEqual(context['published_result_count'], 2)
        # Semester 1's CA is gone, but its result is not.
        self.assertNotIn('PHM101', [m['module_code'] for m in context['ca_theory_modules']])


class CompulsoryFormTests(FormTestBase):
    """A form the college declares compulsory holds the portal until it is
    answered — the same block an unchanged password already puts in the way.

    Response rates on a form students can ignore are what make an evaluation
    useless to report on, so this is the college saying "everyone, not the
    self-selecting few".
    """

    def _required_form(self, **kw):
        form = self._form(slug='census', title='Course Evaluation', is_mandatory=True, **kw)
        question = self._question(required=True, text='Rate the tutor')
        return form, question

    def _dashboard(self):
        return self.portal.get('/student-dashboard/')

    def test_the_portal_stops_at_a_compulsory_form(self):
        form, question = self._required_form()
        response = self._dashboard()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'student_required_form.html')
        self.assertContains(response, 'Course Evaluation')
        self.assertContains(response, 'every student to complete this')
        # None of the portal is reachable behind it.
        self.assertNotContains(response, 'data-view="obligations"')

    def test_answering_it_releases_the_portal(self):
        form, question = self._required_form()
        self.assertTemplateUsed(self._dashboard(), 'student_required_form.html')

        submitted = self.portal.post(
            f'/api/my-forms/{form.slug}/submit/',
            {'answers': {str(question.id): 'Good'}}, content_type='application/json')
        self.assertEqual(submitted.status_code, 201, submitted.content)

        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')

    def test_an_optional_form_never_blocks(self):
        self._form(slug='optional-one', is_mandatory=False)
        self._question()
        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')

    def test_a_compulsory_form_that_is_not_live_does_not_block(self):
        """Drafting one must not lock every student out before it is ready."""
        self._form(slug='draft-census', is_mandatory=True, is_active=False)
        self._question()
        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')

    def test_it_stops_blocking_once_the_window_closes(self):
        self._form(slug='closed-census', is_mandatory=True,
                   closes_on=date(2020, 1, 1))
        self._question()
        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')

    def test_forms_are_asked_one_at_a_time(self):
        first, q1 = self._required_form()
        second = self._form(slug='census-two', title='Hostel Evaluation', is_mandatory=True)
        q2 = self._question(required=True, text='Rate the hostel')

        page = self._dashboard()
        self.assertContains(page, 'Course Evaluation')
        self.assertContains(page, 'more form')          # says one is still to come
        self.assertNotContains(page, 'Rate the hostel')

        self.portal.post(f'/api/my-forms/{first.slug}/submit/',
                         {'answers': {str(q1.id): 'Good'}}, content_type='application/json')
        page = self._dashboard()
        self.assertTemplateUsed(page, 'student_required_form.html')
        self.assertContains(page, 'Hostel Evaluation')

        self.portal.post(f'/api/my-forms/{second.slug}/submit/',
                         {'answers': {str(q2.id): 'Good'}}, content_type='application/json')
        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')

    def test_a_compulsory_form_can_still_be_anonymous(self):
        """The college learns everyone answered without learning who said what
        — the receipt clears the block, not the response."""
        form, question = self._required_form(is_anonymous=True)
        self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                         {'answers': {str(question.id): 'Poor'}},
                         content_type='application/json')

        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')
        self.assertIsNone(FormResponse.objects.get(form=form).profile)
        self.assertTrue(FormSubmissionReceipt.objects.filter(
            form=form, profile=self.asha).exists())

    def test_a_staff_form_never_blocks_a_student(self):
        self._form(slug='mentor-census', is_mandatory=True, audience=Form.STAFF)
        self._question()
        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')

    def test_the_student_can_still_sign_out(self):
        """Blocked is not trapped."""
        self._required_form()
        self.assertContains(self._dashboard(), 'Sign out')

    def test_a_required_question_is_still_enforced(self):
        form, question = self._required_form()
        refused = self.portal.post(f'/api/my-forms/{form.slug}/submit/',
                                   {'answers': {}}, content_type='application/json')
        self.assertEqual(refused.status_code, 400)
        self.assertTemplateUsed(self._dashboard(), 'student_required_form.html')

    def test_the_admin_can_declare_and_withdraw_it(self):
        form = self._form(slug='census', is_mandatory=False)
        self._question()
        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')

        self.assertEqual(self.api.patch(f'/api/forms/{form.id}/',
                                        {'is_mandatory': True}, format='json').status_code, 200)
        self.assertTemplateUsed(self._dashboard(), 'student_required_form.html')

        self.api.patch(f'/api/forms/{form.id}/', {'is_mandatory': False}, format='json')
        self.assertTemplateUsed(self._dashboard(), 'student_dashboard.html')


class LevelTargetedFormsTests(FormTestBase):
    """The college does not ask every NTA level the same things.

    The introduction letter request exists in two versions — one for levels 4
    and 5, another for level 6 — and a student who fills in the wrong one has
    given the college the wrong answers, not merely their own time.
    """

    def setUp(self):
        super().setUp()
        # setUp's own student, Asha, is at level 4. Add a level 6 student so
        # both sides of a targeted form can be checked.
        self.level6 = ClassLevel.objects.create(name='NTA Level 6', order=6)
        module6 = Module.objects.create(name='Pharmacotherapy', code='PHM301', teacher='T',
                                        class_level=self.level6, semester=self.sem)
        enrollment = Student.objects.create(
            nactvet_reg_no='BPH/2026/006', name='Neema Paul', module=module6)
        enrollment.set_portal_pin('Portal#2026', require_change=False)
        enrollment.save()
        from . import finance
        self.neema_profile = finance.profile_for_student(enrollment)
        self.neema = Client()
        self.neema.post('/login/', {'identifier': 'BPH/2026/006', 'secret': 'Portal#2026'})

    def _slugs(self, portal):
        return sorted(f['slug'] for f in portal.get('/api/my-forms/').json())

    def test_a_form_with_no_levels_goes_to_everyone(self):
        self._form(title='Course Evaluation', slug='course-eval')
        self.assertEqual(self._slugs(self.portal), ['course-eval'])
        self.assertEqual(self._slugs(self.neema), ['course-eval'])

    def test_a_level_6_form_never_reaches_a_level_4_student(self):
        form = self._form(title='Letter request (level 6)', slug='letter-6')
        form.levels.set([self.level6])

        self.assertEqual(self._slugs(self.neema), ['letter-6'])
        self.assertEqual(self._slugs(self.portal), [])

    def test_a_form_can_be_aimed_at_two_levels(self):
        level5 = ClassLevel.objects.create(name='NTA Level 5', order=5)
        form = self._form(title='Letter request (levels 4 and 5)', slug='letter-45')
        form.levels.set([self.level, level5])

        self.assertEqual(self._slugs(self.portal), ['letter-45'])
        self.assertEqual(self._slugs(self.neema), [])

    def test_knowing_the_slug_is_not_a_way_in(self):
        """Hidden from the list is not enough — the endpoints refuse it too, and
        say the same thing they would about a closed form."""
        form = self._form(title='Letter request (level 6)', slug='letter-6')
        form.levels.set([self.level6])
        self._question(required=True)

        opened = self.portal.get('/api/my-forms/letter-6/')
        self.assertEqual(opened.status_code, 404)
        self.assertEqual(opened.json()['detail'], 'That form is not open for responses.')

        sent = self.portal.post(
            '/api/my-forms/letter-6/submit/',
            {'answers': {str(FormQuestion.objects.first().id): 'Good'}},
            content_type='application/json')
        self.assertEqual(sent.status_code, 400)
        self.assertEqual(sent.json()['detail'], 'That form is not open for responses.')
        self.assertEqual(FormResponse.objects.count(), 0)

    def test_the_student_it_is_aimed_at_can_still_answer(self):
        form = self._form(title='Letter request (level 6)', slug='letter-6')
        form.levels.set([self.level6])
        question = self._question(required=True)

        sent = self.neema.post(
            '/api/my-forms/letter-6/submit/',
            {'answers': {str(question.id): 'Good'}}, content_type='application/json')
        self.assertEqual(sent.status_code, 201)
        self.assertEqual(FormResponse.objects.get().profile, self.neema_profile)

    def test_a_compulsory_form_for_another_level_does_not_block(self):
        form = self._form(title='Level 6 census', slug='census-6', is_mandatory=True)
        form.levels.set([self.level6])
        self._question(required=True)

        held = self.neema.get('/student-dashboard/')
        self.assertTemplateUsed(held, 'student_required_form.html')

        free = self.portal.get('/student-dashboard/')
        self.assertTemplateUsed(free, 'student_dashboard.html')

    def test_a_student_with_no_enrollment_is_not_shown_a_targeted_form(self):
        """We cannot tell what level they are at, and guessing puts the level 6
        letter request in front of a level 4 student."""
        form = self._form(title='Letter request (level 6)', slug='letter-6')
        form.levels.set([self.level6])
        stray = self.asha.enrollments.all()
        stray.delete()

        self.assertEqual(evaluations.forms_for_student(self.asha), [])
        self.assertEqual(evaluations.pending_mandatory_forms(self.asha), [])

    def test_the_admin_can_aim_a_form_and_change_its_mind(self):
        created = self.api.post('/api/forms/', {
            'title': 'Introduction letter request (level 6)',
            'levels': [self.level6.id],
        }, format='json')
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.data['levels'], [self.level6.id])
        self.assertEqual(created.data['level_names'], 'NTA Level 6')

        form = Form.objects.get(id=created.data['id'])
        self.assertEqual([level.name for level in form.levels.all()], ['NTA Level 6'])

        widened = self.api.patch(f'/api/forms/{form.id}/',
                                 {'levels': [self.level.id, self.level6.id]}, format='json')
        self.assertEqual(widened.status_code, 200)
        self.assertEqual(widened.data['level_names'], 'NTA Level 4, NTA Level 6')

        released = self.api.patch(f'/api/forms/{form.id}/', {'levels': []}, format='json')
        self.assertEqual(released.data['level_names'], 'All levels')
        self.assertEqual(self._slugs(self.portal), [])   # still a draft, so still hidden

    def test_the_seeded_letter_requests_land_at_the_right_levels(self):
        ClassLevel.objects.create(name='NTA Level 5', order=5)
        call_command('seed_evaluation_forms', verbosity=0)

        by_slug = {form.slug: form for form in Form.objects.all()}
        self.assertEqual(
            sorted(level.order for level in by_slug['practicum-letter-request'].levels.all()),
            [4, 5])
        self.assertEqual(
            [level.order for level in by_slug['practicum-letter-request-level-6'].levels.all()],
            [6])
        # Everything else stays open to every level.
        self.assertEqual(list(by_slug['course-evaluation'].levels.all()), [])

    def test_a_continuing_student_gets_this_years_level(self):
        """Asha was a level 4 last year and is a level 5 now. The level 5 form
        is the one she is asked for, and last year's is not."""
        level5 = ClassLevel.objects.create(name='NTA Level 5', order=5)
        last_year = AcademicYear.objects.create(name='2025/2026', is_active=False)
        old_sem = Semester.objects.create(academic_year=last_year, number=1)
        self.enrollment.module = Module.objects.create(
            name='Anatomy', code='ANA101', teacher='T',
            class_level=self.level, semester=old_sem)
        self.enrollment.save()
        this_year = Module.objects.create(name='Dispensing', code='DSP201', teacher='T',
                                          class_level=level5, semester=self.sem)
        Student.objects.create(nactvet_reg_no='BPH/2026/001', name='Asha Juma',
                               module=this_year, profile=self.asha)

        for_five = self._form(title='Level 5 form', slug='level-5-form')
        for_five.levels.set([level5])
        for_four = self._form(title='Level 4 form', slug='level-4-form')
        for_four.levels.set([self.level])

        self.assertEqual(self._slugs(self.portal), ['level-5-form'])


class ServiceRequestTests(FormTestBase):
    """A service request is one student asking the college for one thing —
    a sick sheet, a letter, a bed — and waiting on an answer.

    It uses the same builder as an evaluation, but the two are not the same
    animal: an evaluation is feedback nobody owes a reply to, and a request is
    owed one, by name.
    """

    def _service(self, **kw):
        defaults = {'title': 'Sick Sheet Request', 'slug': 'sick-sheet',
                    'kind': Form.REQUEST, 'allow_multiple': True}
        defaults.update(kw)
        form = self._form(**defaults)
        self.reason = self._question(text='What is wrong', type=FormQuestion.SHORT_TEXT,
                                     options=[], required=True)
        return form

    def _ask(self, portal=None, answer='Malaria'):
        return (portal or self.portal).post(
            f'/api/my-forms/{self.reason.section.form.slug}/submit/',
            {'answers': {str(self.reason.id): answer}}, content_type='application/json')

    # ── the two lists are not the same list ──────────────────────────────────

    def test_a_service_is_kept_out_of_the_evaluation_list(self):
        self._service()
        self._form(title='Course Evaluation', slug='course-eval')   # an evaluation

        page = self.portal.get('/student-dashboard/')
        self.assertEqual([e['form'].slug for e in page.context['student_forms']],
                         ['course-eval'])
        self.assertEqual([e['form'].slug for e in page.context['service_forms']],
                         ['sick-sheet'])

    def test_the_portal_offers_services_under_their_own_menu(self):
        self._service()
        page = self.portal.get('/student-dashboard/')
        self.assertContains(page, 'data-view="services"')
        self.assertContains(page, 'data-view="my-requests"')
        # Forms moved in beside them rather than staying a menu of its own.
        self.assertContains(page, 'Request a Service')

    # ── asking, and being answered ───────────────────────────────────────────

    def test_a_request_arrives_with_the_college_and_the_student_can_see_it(self):
        self._service()
        self.assertEqual(self._ask().status_code, 201)

        made = FormResponse.objects.get()
        self.assertEqual(made.profile, self.asha)
        self.assertEqual(made.status, FormResponse.PENDING)
        self.assertIsNone(made.decided_at)

        page = self.portal.get('/student-dashboard/')
        self.assertEqual(list(page.context['my_requests']), [made])
        self.assertEqual(page.context['requests_pending'], 1)
        self.assertContains(page, 'With the college')

    def test_the_admin_answers_it_and_the_student_is_told(self):
        self._service()
        self._ask()
        made = FormResponse.objects.get()

        queue = self.api.get('/api/service-requests/?status=pending')
        self.assertEqual([row['id'] for row in queue.data], [made.id])
        self.assertEqual(queue.data[0]['student_name'], 'Asha Juma')

        answered = self.api.post(f'/api/service-requests/{made.id}/decide/', {
            'status': 'approved',
            'note': 'Collect it from the Dean of Students on Tuesday.',
        }, format='json')
        self.assertEqual(answered.status_code, 200, answered.data)

        made.refresh_from_db()
        self.assertEqual(made.status, FormResponse.APPROVED)
        self.assertEqual(made.decided_by, self.admin)
        self.assertIsNotNone(made.decided_at)

        page = self.portal.get('/student-dashboard/')
        self.assertContains(page, 'Collect it from the Dean of Students on Tuesday.')
        self.assertEqual(page.context['requests_pending'], 0)

    def test_declining_says_why(self):
        self._service()
        self._ask()
        made = FormResponse.objects.get()
        self.api.post(f'/api/service-requests/{made.id}/decide/', {
            'status': 'declined', 'note': 'You have already been issued one this month.',
        }, format='json')

        page = self.portal.get('/student-dashboard/')
        self.assertContains(page, 'Declined')
        self.assertContains(page, 'already been issued one this month')

    def test_a_decision_can_be_undone(self):
        """An officer who approved the wrong request should not have to live
        with a decision nobody made."""
        self._service()
        self._ask()
        made = FormResponse.objects.get()
        self.api.post(f'/api/service-requests/{made.id}/decide/',
                      {'status': 'approved', 'note': 'Ready'}, format='json')
        self.api.post(f'/api/service-requests/{made.id}/decide/',
                      {'status': 'pending'}, format='json')

        made.refresh_from_db()
        self.assertEqual(made.status, FormResponse.PENDING)
        self.assertIsNone(made.decided_by)
        self.assertIsNone(made.decided_at)

    def test_only_the_admin_answers_a_request(self):
        self._service()
        self._ask()
        made = FormResponse.objects.get()

        tutor = User.objects.create_user('tutor', password='pw')
        TeacherProfile.objects.create(user=tutor, full_name='Tutor')
        api = APIClient()
        api.force_authenticate(tutor)

        refused = api.post(f'/api/service-requests/{made.id}/decide/',
                           {'status': 'approved'}, format='json')
        self.assertEqual(refused.status_code, 403)
        made.refresh_from_db()
        self.assertEqual(made.status, FormResponse.PENDING)
        # A tutor may still read the queue.
        self.assertEqual(api.get('/api/service-requests/').status_code, 200)

    def test_an_evaluation_is_not_something_to_approve(self):
        form = self._form(title='Course Evaluation', slug='course-eval')
        question = self._question()
        response = evaluations.submit(form, {question.id: 'Good'}, profile=self.asha)

        refused = self.api.post(f'/api/service-requests/{response.id}/decide/',
                                {'status': 'approved'}, format='json')
        self.assertEqual(refused.status_code, 404)   # not in the request queue at all
        with self.assertRaises(ValueError):
            evaluations.decide(response, status='approved')

    # ── the rules a request has to obey ──────────────────────────────────────

    def test_a_request_can_be_neither_anonymous_nor_compulsory(self):
        for field in ('is_anonymous', 'is_mandatory'):
            refused = self.api.post('/api/forms/', {
                'title': 'Sick Sheet Request', 'kind': 'request', field: True,
            }, format='json')
            self.assertEqual(refused.status_code, 400, field)
            self.assertIn(field, refused.data)

    def test_an_existing_form_cannot_be_turned_into_an_anonymous_request(self):
        form = self._form(title='Tutor Evaluation', slug='tutor-eval', is_anonymous=True)
        refused = self.api.patch(f'/api/forms/{form.id}/', {'kind': 'request'}, format='json')
        self.assertEqual(refused.status_code, 400)
        form.refresh_from_db()
        self.assertEqual(form.kind, Form.EVALUATION)

    def test_a_compulsory_request_still_does_not_block_the_portal(self):
        """The API refuses the combination, so this can only arrive as bad data
        — and bad data must not lock every student behind a sick sheet."""
        form = self._service()
        Form.objects.filter(id=form.id).update(is_mandatory=True)

        page = self.portal.get('/student-dashboard/')
        self.assertTemplateUsed(page, 'student_dashboard.html')

    def test_a_service_is_aimed_at_levels_like_any_other_form(self):
        level6 = ClassLevel.objects.create(name='NTA Level 6', order=6)
        form = self._service(title='Letter request', slug='letter-6')
        form.levels.set([level6])

        page = self.portal.get('/student-dashboard/')
        self.assertEqual(page.context['service_forms'], [])
        self.assertEqual(self._ask().status_code, 400)

    def test_a_service_that_is_not_open_cannot_be_requested(self):
        self._service(is_active=False)
        page = self.portal.get('/student-dashboard/')
        self.assertEqual(page.context['service_forms'], [])
        self.assertEqual(self._ask().status_code, 400)
        self.assertEqual(FormResponse.objects.count(), 0)

    def test_one_bed_one_request(self):
        """A hostel place is asked for once a year, and the portal says so
        rather than letting a student queue twice."""
        self._service(title='Hostel Accommodation Request', slug='hostel-request',
                      allow_multiple=False)
        self.assertEqual(self._ask().status_code, 201)

        page = self.portal.get('/student-dashboard/')
        self.assertFalse(page.context['service_forms'][0]['can_answer'])
        self.assertEqual(self._ask().status_code, 400)
        self.assertEqual(FormResponse.objects.count(), 1)


class SecretaryTests(FormTestBase):
    """The secretary is the office a request for a sick sheet or leave of
    absence goes to, and the person who releases the printed document.

    A separate role because the work is separate: they have no business in the
    ledger or the academic register, and the accountant has none in theirs.
    """

    def setUp(self):
        super().setUp()
        self.secretary = User.objects.create_user('secretary', password='pw')
        SecretaryProfile.objects.create(user=self.secretary, full_name='Neema Katabazi')
        self.sec_api = APIClient()
        self.sec_api.force_authenticate(self.secretary)

        self.service = self._form(title="Student's Sicksheet Form", slug='sick-sheet',
                                  kind=Form.REQUEST, allow_multiple=True)
        self.facility = self._question(text='Name of the health facility',
                                       type=FormQuestion.SHORT_TEXT, options=[], required=True)

    def _ask(self, answer='Singida Regional Referral Hospital'):
        return self.portal.post('/api/my-forms/sick-sheet/submit/',
                                {'answers': {str(self.facility.id): answer}},
                                content_type='application/json')

    def test_the_admin_can_create_a_secretary_account(self):
        made = self.api.post('/api/staff-accounts/', {
            'role': 'secretary', 'full_name': 'Asha Mtei',
            'username': 'amtei', 'password': 'secret123',
        }, format='json')
        self.assertEqual(made.status_code, 201, made.data)
        self.assertEqual(made.data['role'], 'secretary')
        self.assertTrue(SecretaryProfile.objects.filter(user__username='amtei').exists())

        listed = self.api.get('/api/staff-accounts/').json()
        self.assertIn(('amtei', 'secretary'),
                      [(row['username'], row['role']) for row in listed])

    def test_the_secretary_answers_requests(self):
        self._ask()
        made = FormResponse.objects.get()

        queue = self.sec_api.get('/api/service-requests/?status=pending')
        self.assertEqual([row['id'] for row in queue.data], [made.id])

        answered = self.sec_api.post(f'/api/service-requests/{made.id}/decide/', {
            'status': 'approved', 'note': 'Print it and take it to the facility.',
        }, format='json')
        self.assertEqual(answered.status_code, 200, answered.data)
        made.refresh_from_db()
        self.assertEqual(made.status, FormResponse.APPROVED)
        self.assertEqual(made.decided_by, self.secretary)

    def test_the_secretary_is_told_apart_from_the_other_roles(self):
        self.sec_api.force_authenticate(self.secretary)
        payload = self.sec_api.get('/api/dashboard/').json()
        self.assertTrue(payload['is_secretary'])
        self.assertFalse(payload['is_staff'])
        self.assertFalse(payload['is_accountant'])
        self.assertFalse(payload['is_estate_officer'])

    def test_a_tutor_still_cannot_answer_a_request(self):
        self._ask()
        made = FormResponse.objects.get()
        tutor = User.objects.create_user('tutor', password='pw')
        TeacherProfile.objects.create(user=tutor, full_name='Tutor')
        api = APIClient()
        api.force_authenticate(tutor)

        self.assertEqual(
            api.post(f'/api/service-requests/{made.id}/decide/',
                     {'status': 'approved'}, format='json').status_code, 403)


class PrintedRequestTests(FormTestBase):
    """An approved request comes back as the college's own paper form, ready to
    be signed and stamped. Nothing else prints: a pending request is not a
    document yet, and a declined one is not a document at all."""

    def setUp(self):
        super().setUp()
        self.service = self._form(title="Student's Sicksheet Form", slug='sick-sheet',
                                  kind=Form.REQUEST, allow_multiple=True,
                                  print_note='Must be stamped by the health facility.')
        self.facility = self._question(text='Name of the health facility',
                                       type=FormQuestion.SHORT_TEXT, options=[], required=True)
        # Part C of the paper form: nobody in the portal fills this in.
        self.office = FormSection.objects.create(
            form=self.service, title='C. APPROVED BY COLLEGE ADMINISTRATION',
            order=1, for_office=True)
        self.signature = FormQuestion.objects.create(
            section=self.office, text='Name of Head of Department',
            type=FormQuestion.SHORT_TEXT, order=0)

    def _approved(self):
        self.portal.post('/api/my-forms/sick-sheet/submit/',
                         {'answers': {str(self.facility.id): 'Singida Referral'}},
                         content_type='application/json')
        made = FormResponse.objects.get()
        evaluations.decide(made, status=FormResponse.APPROVED, note='Collect it today.',
                           by=self.admin)
        return made

    def test_an_office_part_is_never_shown_to_the_student(self):
        opened = self.portal.get('/api/my-forms/sick-sheet/').json()
        titles = [section['title'] for section in opened['sections']]
        self.assertEqual(titles, ['Section A'])
        # And an answer to one is not accepted either.
        refused = self.portal.post('/api/my-forms/sick-sheet/submit/', {
            'answers': {str(self.facility.id): 'X', str(self.signature.id): 'Me'},
        }, content_type='application/json')
        self.assertEqual(refused.status_code, 400)

    def test_an_office_part_is_never_required_of_the_student(self):
        FormQuestion.objects.filter(id=self.signature.id).update(required=True)
        sent = self.portal.post('/api/my-forms/sick-sheet/submit/',
                                {'answers': {str(self.facility.id): 'Singida Referral'}},
                                content_type='application/json')
        self.assertEqual(sent.status_code, 201)

    def test_an_approved_request_prints_on_the_college_letterhead(self):
        made = self._approved()
        page = self.portal.get(f'/request/{made.id}/')

        self.assertEqual(page.status_code, 200)
        self.assertTemplateUsed(page, 'request_print.html')
        self.assertContains(page, made.reference)
        self.assertContains(page, 'Asha Juma')
        self.assertContains(page, 'BPH/2026/001')
        self.assertContains(page, 'Singida Referral')
        # The office part is printed, blank, for whoever signs it.
        self.assertContains(page, 'C. APPROVED BY COLLEGE ADMINISTRATION')
        self.assertContains(page, 'Name of Head of Department')
        self.assertContains(page, 'Official stamp')
        self.assertContains(page, 'Must be stamped by the health facility.')

    def test_download_goes_straight_to_save_as_pdf(self):
        made = self._approved()
        self.assertContains(self.portal.get(f'/request/{made.id}/?download=1'), 'window.print()')

    def test_nothing_prints_until_it_is_approved(self):
        self.portal.post('/api/my-forms/sick-sheet/submit/',
                         {'answers': {str(self.facility.id): 'Singida Referral'}},
                         content_type='application/json')
        made = FormResponse.objects.get()
        self.assertEqual(self.portal.get(f'/request/{made.id}/').status_code, 404)

        evaluations.decide(made, status=FormResponse.DECLINED, note='No.', by=self.admin)
        self.assertEqual(self.portal.get(f'/request/{made.id}/').status_code, 404)

    def test_a_student_can_only_print_their_own(self):
        made = self._approved()
        other_module = Module.objects.create(name='Anatomy', code='ANA101', teacher='T',
                                             class_level=self.level, semester=self.sem)
        other = Student.objects.create(nactvet_reg_no='BPH/2026/999', name='Someone Else',
                                       module=other_module)
        other.set_portal_pin('Portal#2026', require_change=False)
        other.save()
        intruder = Client()
        intruder.post('/login/', {'identifier': 'BPH/2026/999', 'secret': 'Portal#2026'})

        self.assertEqual(intruder.get(f'/request/{made.id}/').status_code, 404)

    def test_the_secretary_can_see_what_the_student_prints(self):
        made = self._approved()
        secretary = User.objects.create_user('secretary', password='pw')
        SecretaryProfile.objects.create(user=secretary, full_name='Neema')
        office = Client()
        office.force_login(secretary)
        self.assertEqual(office.get(f'/request/{made.id}/').status_code, 200)

    def test_a_stranger_is_sent_to_the_login_screen(self):
        made = self._approved()
        self.assertEqual(Client().get(f'/request/{made.id}/').status_code, 302)


class SeededCollegeFormsTests(TestCase):
    """The sick sheet and the permission form, transcribed from the college's
    own paper documents."""

    def test_the_sick_sheet_matches_the_paper_form(self):
        call_command('seed_evaluation_forms', verbosity=0)
        form = Form.objects.get(slug='sick-sheet-request')

        self.assertEqual(form.kind, Form.REQUEST)
        self.assertFalse(form.is_anonymous)
        self.assertIn('attendance percentage (90%)', form.print_note)

        sections = list(form.sections.all())
        # Only Part A's facility details are the student's to fill in; the rest
        # of the paper form belongs to the requesting officer, the health
        # facility and the college administration.
        student_parts = [s.title for s in sections if not s.for_office]
        self.assertEqual(student_parts, ['A. REQUEST TO HEALTH FACILITY'])
        office_parts = [s.title for s in sections if s.for_office]
        self.assertEqual(len(office_parts), 5)
        self.assertTrue(any('Head of Department' in t for t in office_parts))
        self.assertTrue(any('Dean of Students' in t for t in office_parts))
        self.assertTrue(any('HEALTH FACILITY DECLARATION' in t for t in office_parts))

    def test_the_permission_form_keeps_the_colleges_own_words(self):
        call_command('seed_evaluation_forms', verbosity=0)
        form = Form.objects.get(slug='student-permission-request')

        self.assertEqual(form.title, 'Fomu ya Ruhusa ya Wanafunzi')
        self.assertEqual(form.kind, Form.REQUEST)

        sections = {s.title: s for s in form.sections.all()}
        self.assertIn('A. TAARIFA ZA MWANAFUNZI', sections)
        self.assertFalse(sections['A. TAARIFA ZA MWANAFUNZI'].for_office)
        # The declaration is the student's to accept; the signature, the Head of
        # Department's comments and the Dean's are all signed on paper.
        self.assertIn('asilimia 90%', sections['B. TAMKO'].description)
        self.assertTrue(sections['C. IMEPITISHWA NA — i. Mkuu wa Idara'].for_office)
        self.assertTrue(sections['C. IMEPITISHWA NA — ii. Muadili wa wanafunzi'].for_office)

        asked = [q.text for q in evaluations.answerable_questions(form)]
        self.assertIn('Sababu ya kutokuwepo Chuoni kwa muda huo', asked)
        self.assertNotIn('Saini', asked)
